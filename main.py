from dotenv import load_dotenv

load_dotenv()

import os
import base64
import json
import io
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from openai import OpenAI
from openpyxl import Workbook
from datetime import datetime
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient

app = FastAPI()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
endpoint = os.getenv("AZURE_DOC_INTELLIGENCE_URL")
key = os.getenv("AZURE_DOC_INTELLIGENCE_KEY")

document_intelligence_client = DocumentIntelligenceClient(
    endpoint=endpoint, credential=AzureKeyCredential(key)
)

PROMPT = """
Olvasd be a feltöltött PDF dokumentum teljes tartalmát, akkor is, ha több oldalas a dokumentum.

A dokumentum egy számla. A feladatod, hogy kinyerd belőle az alábbi adatokat:

- számlaszám
- vevő neve
- szállító neve
- vevő adószáma
- szállító adószáma
- teljesítés dátuma
- számla kelte
- fizetési határidő
- bruttó összeg (összesen)
- nettó összeg (összesen)
- áfa összege (összesen)
- devizanem

Ezen felül, ha a számlán több tétel szerepel, azokat is add vissza egy 'tetelek' nevű tömbben.
Minden tétel egy objektum legyen a következő szerkezettel:
{
  "megnevezes": "",
  "netto": "",
  "afa": "",
  "afakulcs": "",
  "brutto": ""
}

❗Fontos szabály:
Az összegekhez (például bruttó, nettó, áfa) **soha ne írd hozzá a devizanemet** (például HUF, EUR, Ft, € stb.).
Az összegek értéke csak a szám legyen, formázás és devizajel nélkül (például "5730.88" vagy "573088").
A devizanem külön mezőben szerepeljen a "devizanem" kulcs alatt.

A választ **csak jól formázott JSON** formátumban add meg az alábbi szerkezet szerint:

{
  "szamlaszam": "",
  "vevo_neve": "",
  "szallito_neve": "",
  "vevo_adoszam": "",
  "szallito_adoszam": "",
  "teljesites_datuma": "",
  "szamla_keltee": "",
  "fizetesi_hatarido": "",
  "brutto_osszeg": "",
  "netto_osszeg": "",
  "afa_osszeg": "",
  "devizanem": "",
  "tetelek": [
    {
      "megnevezes": "",
      "netto": "",
      "afa": "",
      "afakulcs": "",
      "brutto": ""
    }
  ]
}

Ha egy adat vagy tétel bármely mezője nem található, az értéke legyen üres string ("").
Ne adj vissza semmit, csak a JSON-t.
"""


@app.post("/import/gpt-5")
async def import_invoice(files: list[UploadFile] = File(...)):

    if not files:
        raise HTTPException(status_code=400, detail="Legalább egy PDF fájlt tölts fel!")

    wb = Workbook()
    ws_invoices = wb.active
    ws_invoices.title = "Számlák"
    ws_items = wb.create_sheet("Tételek")

    # Fejlécek
    invoice_headers = [
        "szamlaszam",
        "vevo_neve",
        "szallito_neve",
        "vevo_adoszam",
        "szallito_adoszam",
        "teljesites_datuma",
        "szamla_keltee",
        "fizetesi_hatarido",
        "brutto_osszeg",
        "netto_osszeg",
        "afa_osszeg",
        "devizanem",
    ]
    ws_invoices.append(invoice_headers)

    item_headers = ["szamlaszam", "megnevezes", "netto", "afa", "afakulcs", "brutto"]
    ws_items.append(item_headers)

    # Feldolgozás minden fájlra
    for file in files:
        if file.content_type != "application/pdf":
            raise HTTPException(
                status_code=400, detail=f"{file.filename} nem PDF fájl!"
            )

        file_bytes = await file.read()
        base64_string = base64.b64encode(file_bytes).decode("utf-8")

        # OpenAI feldolgozás
        response = client.responses.create(
            model="gpt-5",
            input=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_file",
                            "filename": file.filename,
                            "file_data": f"data:application/pdf;base64,{base64_string}",
                        },
                        {"type": "input_text", "text": PROMPT},
                    ],
                },
            ],
        )

        text = response.output_text.strip()

        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            raise HTTPException(
                status_code=500,
                detail=f"{file.filename} feldolgozása sikertelen, nem érvényes JSON.",
            )

        # Fejlécadatok hozzáadása a "Számlák" munkalaphoz
        ws_invoices.append([data.get(h, "") for h in invoice_headers])

        # Tételek hozzáadása a "Tételek" munkalaphoz
        szamlaszam = data.get("szamlaszam", "")
        for item in data.get("tetelek", []):
            ws_items.append(
                [
                    szamlaszam,
                    item.get("megnevezes", ""),
                    item.get("netto", ""),
                    item.get("afa", ""),
                    item.get("afakulcs", ""),
                    item.get("brutto", ""),
                ]
            )

    # Excel mentése memóriába
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    filename = f"invoices_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/import/azure-ai")
async def import_invoice(file: UploadFile = File(...)):
    file_bytes = await file.read()

    # Base64 kódolás az Azure API-hoz
    base64_encoded = base64.b64encode(file_bytes).decode("utf-8")

    poller = document_intelligence_client.begin_analyze_document(
        model_id="prebuilt-invoice",
        body={"base64Source": base64_encoded},
    )

    result = poller.result()

    wb = Workbook()
    ws_headers = wb.active
    ws_headers.title = "Számlák"
    ws_items = wb.create_sheet("Tételek")
    # ws_tables = wb.create_sheet("Tételek (tábla)")

    invoice_headers_written = False
    item_headers_written = False
    table_headers_written = False

    for invoice in result.documents:
        fields = invoice.fields or {}

        # Számla mezők dinamikusan
        invoice_data = {}
        for key, field in fields.items():
            value = (
                field.value
                if hasattr(field, "value")
                else field.content if hasattr(field, "content") else ""
            )
            invoice_data[key] = str(value)

        if not invoice_headers_written:
            ws_headers.append(list(invoice_data.keys()))
            invoice_headers_written = True

        ws_headers.append(list(invoice_data.values()))

        # Tételmezők (Items)
        items = fields.get("Items")
        if items and items.value_array:
            for item in items.value_array:
                item_obj = item.value_object
                item_data = {}
                for k, v in item_obj.items():
                    val = (
                        v.value
                        if hasattr(v, "value")
                        else v.content if hasattr(v, "content") else ""
                    )
                    item_data[k] = str(val)

                # Első alkalommal: oszlopnevek
                if not item_headers_written:
                    ws_items.append(["InvoiceId"] + list(item_data.keys()))
                    item_headers_written = True

                ws_items.append(
                    [invoice_data.get("InvoiceId", "")] + list(item_data.values())
                )

    # # 🆕 Táblázatos tételek beolvasása a "tables" kulcsból
    # if result.tables:
    #     for table in result.tables:
    #         rows = table.row_count
    #         cols = table.column_count
    #         cells = table.cells

    #         # Cellák mátrixba
    #         matrix = [["" for _ in range(cols)] for _ in range(rows)]
    #         for cell in cells:
    #             r = cell.row_index
    #             c = cell.column_index
    #             matrix[r][c] = cell.content

    #         # Fejléc
    #         headers = matrix[0]

    #         if not table_headers_written:
    #             ws_tables.append(headers)
    #             table_headers_written = True

    #         for row in matrix[1:]:
    #             ws_tables.append(row)

    if result.tables:
        # 🆕 Összes táblázat egymás után egy munkalapon ("Összes tábla")
        ws_all_tables = wb.create_sheet("Összes tábla")
        table_headers_written = False

        for table in result.tables:
            rows = table.row_count
            cols = table.column_count
            cells = table.cells

            # Cellák mátrixba szervezése
            matrix = [["" for _ in range(cols)] for _ in range(rows)]
            for cell in cells:
                r = cell.row_index
                c = cell.column_index
                matrix[r][c] = cell.content

            # Fejléc: első sor a táblában
            headers = matrix[0]

            # Egyszer írjuk ki a fejlécet, az első tábla alapján
            if not table_headers_written:
                ws_all_tables.append(headers)
                table_headers_written = True

            # Adatsorokat hozzáadni (fejlécen kívüli sorokat)
            for row in matrix[1:]:
                ws_all_tables.append(row)

            # Üres sor táblák közé elválasztónak (opcionális)
            ws_all_tables.append([])

    # Excel mentése
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"invoice_data_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
